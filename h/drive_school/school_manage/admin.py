from django.contrib import admin
from django.utils.html import format_html
from .models import Coach, Student, CoachSettlement
import openpyxl
from django.http import HttpResponse

@admin.register(Coach)
class CoachAdmin(admin.ModelAdmin):
    list_display = ['name', 'phone', 'car_plate', 'created_at']
    search_fields = ['name', 'phone', 'car_plate']
    list_filter = ['created_at']
    actions = ['export_to_excel']
    
    def get_actions(self, request):
        actions = super().get_actions(request)
        if 'actions' in actions:
            del actions['actions']
        return actions
    
    def export_to_excel(self, request, queryset):
        """导出教练数据到Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "教练数据"
        
        # 表头
        headers = ['教练姓名', '联系电话', '车牌号', '创建时间', '更新时间']
        ws.append(headers)
        
        # 数据
        for coach in queryset:
            ws.append([
                coach.name,
                coach.phone,
                coach.car_plate or '',
                coach.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                coach.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        # 响应
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=coaches.xlsx'
        wb.save(response)
        return response
    
    export_to_excel.short_description = "导出选中教练到Excel"

@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    list_display = [
        'student_id', 
        'name', 
        'phone', 
        'car_type', 
        'coach', 
        'status', 
        'prepay_coach_amount_display',
        'prepay_coach_date_display',
        'second_pay_coach_amount_display',
        'second_pay_coach_date_display',
        'settlement_completed', 
        'register_date'
    ]
    
    list_filter = ['status', 'car_type', 'coach', 'register_date', 'settlement_completed']
    search_fields = ['student_id', 'name', 'phone', 'id_card']
    readonly_fields = ['student_id', 'created_at', 'updated_at']
    list_per_page = 20
    actions = ['export_to_excel']
    
    def get_actions(self, request):
        actions = super().get_actions(request)
        if 'actions' in actions:
            del actions['actions']
        return actions
    
    def prepay_coach_amount_display(self, obj):
        """显示预付教练金额"""
        if obj.prepay_coach_amount:
            return f"¥{obj.prepay_coach_amount}"
        return "未支付"
    prepay_coach_amount_display.short_description = '预付教练金额'
    prepay_coach_amount_display.admin_order_field = 'prepay_coach_amount'
    
    def prepay_coach_date_display(self, obj):
        """显示预付教练时间"""
        if obj.prepay_coach_date:
            return obj.prepay_coach_date.strftime('%Y-%m-%d')
        return "未设置"
    prepay_coach_date_display.short_description = '预付时间'
    prepay_coach_date_display.admin_order_field = 'prepay_coach_date'
    
    def second_pay_coach_amount_display(self, obj):
        """显示第二次支付教练金额"""
        if obj.second_pay_coach_amount:
            return f"¥{obj.second_pay_coach_amount}"
        return "未支付"
    second_pay_coach_amount_display.short_description = '第二次支付金额'
    second_pay_coach_amount_display.admin_order_field = 'second_pay_coach_amount'
    
    def second_pay_coach_date_display(self, obj):
        """显示第二次支付教练时间"""
        if obj.second_pay_coach_date:
            return obj.second_pay_coach_date.strftime('%Y-%m-%d')
        return "未设置"
    second_pay_coach_date_display.short_description = '第二次支付时间'
    second_pay_coach_date_display.admin_order_field = 'second_pay_coach_date'
    
    fieldsets = (
        ('基本信息', {
            'fields': ('student_id', 'name', 'phone', 'id_card', 'register_date', 'car_type', 'coach', 'status')
        }),
        ('费用信息', {
            'fields': ('total_fee',)
        }),
        ('资金记录', {
            'fields': (
                ('registration_fee', 'registration_fee_date'),
                ('tuition_cash', 'tuition_cash_date'),
            ),
            'description': '填写报名费和学费收取情况'
        }),
        ('时间信息', {
            'fields': ('transfer_date', 'start_training_date', 'graduation_date')
        }),
        ('银行转账记录', {
            'fields': (
                ('bank_transfer1_amount', 'bank_transfer1_date'),
                ('bank_transfer2_amount', 'bank_transfer2_date'),
                ('bank_transfer3_amount', 'bank_transfer3_date'),
            )
        }),
        ('教练支付与结算', {
            'fields': (
                ('prepay_coach_amount', 'prepay_coach_date'),
                ('second_pay_coach_amount', 'second_pay_coach_date'),
                'settlement_completed',
            )
        }),
        ('其他信息', {
            'fields': ('car_number', 'notes')
        }),
    )
    
    def export_to_excel(self, request, queryset):
        """导出Excel功能"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "学员数据"
        
        # 表头
        headers = ['学员编号', '姓名', '手机号', '身份证号', '报名日期', '报考车型', '教练', 
                  '总费用', '状态', '报名费金额', '报名费收取时间', '学费现金金额', '学费现金时间',
                  '预付教练金额', '预付教练时间', '第二次支付教练金额', '第二次支付教练时间',
                  '结算完毕', '备注']
        ws.append(headers)
        
        # 数据
        for student in queryset:
            ws.append([
                student.student_id,
                student.name,
                student.phone,
                student.id_card,
                student.register_date,
                student.get_car_type_display(),
                str(student.coach) if student.coach else '',
                student.total_fee,
                student.get_status_display(),
                student.registration_fee,
                student.registration_fee_date,
                student.tuition_cash,
                student.tuition_cash_date,
                student.prepay_coach_amount,
                student.prepay_coach_date,
                student.second_pay_coach_amount,
                student.second_pay_coach_date,
                student.get_settlement_completed_display(),
                student.notes
            ])
        
        # 响应
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students.xlsx'
        wb.save(response)
        return response
    
    export_to_excel.short_description = "导出选中学员到Excel"

@admin.register(CoachSettlement)
class CoachSettlementAdmin(admin.ModelAdmin):
    list_display = [
        'coach', 
        'settlement_month', 
        'prepay_amount', 
        'prepay_date_display',
        'total_amount', 
        'is_paid_display', 
        'created_at_display'
    ]
    
    list_filter = ['settlement_month', 'is_paid']
    search_fields = ['coach__name']
    filter_horizontal = ['graduated_students']
    actions = ['export_to_excel']
    
    def get_actions(self, request):
        actions = super().get_actions(request)
        if 'actions' in actions:
            del actions['actions']
        return actions
    
    def prepay_date_display(self, obj):
        """显示预付教练时间"""
        if obj.prepay_date:
            return obj.prepay_date.strftime('%Y-%m-%d')
        return "未设置"
    prepay_date_display.short_description = '预付时间'
    prepay_date_display.admin_order_field = 'prepay_date'
    
    def is_paid_display(self, obj):
        """显示是否已支付"""
        if obj.is_paid:
            return format_html('<span style="color: green; font-weight: bold;">✓ 已支付</span>')
        return format_html('<span style="color: red;">未支付</span>')
    is_paid_display.short_description = '支付状态'
    
    def created_at_display(self, obj):
        """格式化创建时间"""
        return obj.created_at.strftime('%Y-%m-%d')
    created_at_display.short_description = '创建日期'
    created_at_display.admin_order_field = 'created_at'
    
    fieldsets = (
        ('基本信息', {
            'fields': ('coach', 'settlement_month', 'graduated_students')
        }),
        ('费用信息', {
            'fields': ('unit_price', 'prepay_amount', 'prepay_date', 'total_amount', 'is_paid')
        }),
    )
    
    def export_to_excel(self, request, queryset):
        """导出教练结算数据到Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "教练结算数据"
        
        # 表头
        headers = ['教练', '结算月份', '单价', '预付教练金额', '预付教练时间', '应结算金额', '是否已支付', '创建时间', '毕业学员数量', '毕业学员名单']
        ws.append(headers)
        
        # 数据
        for settlement in queryset:
            graduated_students = list(settlement.graduated_students.all())
            student_names = '、'.join([student.name for student in graduated_students])
            
            ws.append([
                settlement.coach.name,
                settlement.settlement_month,
                settlement.unit_price,
                settlement.prepay_amount,
                settlement.prepay_date,
                settlement.total_amount,
                '是' if settlement.is_paid else '否',
                settlement.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                len(graduated_students),
                student_names
            ])
        
        # 响应
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=coach_settlements.xlsx'
        wb.save(response)
        return response
    
    export_to_excel.short_description = "导出选中教练结算到Excel"