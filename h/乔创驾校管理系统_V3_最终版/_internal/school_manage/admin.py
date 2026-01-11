# from django.contrib import admin
# from django.utils.html import format_html
# from .models import Coach, Student, CoachSettlement
# import openpyxl
# from django.http import HttpResponse

# @admin.register(Coach)
# class CoachAdmin(admin.ModelAdmin):
#     list_display = ['name', 'phone', 'car_plate', 'created_at']
#     search_fields = ['name', 'phone', 'car_plate']
#     list_filter = ['created_at']

# @admin.register(Student)
# class StudentAdmin(admin.ModelAdmin):
#     list_display = ['student_id', 'name', 'phone', 'car_type', 'coach', 'status', 'total_fee', 'settlement_completed', 'register_date']
#     list_filter = ['status', 'car_type', 'coach', 'register_date', 'settlement_completed']
#     search_fields = ['student_id', 'name', 'phone', 'id_card']
#     readonly_fields = ['student_id', 'created_at', 'updated_at']
#     list_per_page = 20
#     actions = ['export_to_excel']
    
#     fieldsets = (
#         ('基本信息', {
#             'fields': ('student_id', 'name', 'phone', 'id_card', 'register_date', 'car_type', 'coach', 'status')
#         }),
#         ('费用信息', {
#             'fields': ('total_fee',)
#         }),
#         ('资金记录', {
#             'fields': (
#                 ('registration_fee', 'registration_fee_date'),
#                 ('tuition_cash', 'tuition_cash_date'),
#             ),
#             'description': '填写报名费和学费收取情况'
#         }),
#         ('时间信息', {
#             'fields': ('transfer_date', 'start_training_date', 'graduation_date')
#         }),
#         ('银行转账记录', {
#             'fields': (
#                 ('bank_transfer1_amount', 'bank_transfer1_date'),
#                 ('bank_transfer2_amount', 'bank_transfer2_date'),
#                 ('bank_transfer3_amount', 'bank_transfer3_date'),
#             )
#         }),
#         ('教练支付与结算', {
#             'fields': (
#                 ('pay_coach_amount', 'pay_coach_date'),
#                 'settlement_completed',
#             )
#         }),
#         ('其他信息', {
#             'fields': ('car_number', 'notes')
#         }),
#     )
    
#     def export_to_excel(self, request, queryset):
#         """导出Excel功能"""
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "学员数据"
        
#         # 表头
#         headers = ['学员编号', '姓名', '手机号', '身份证号', '报名日期', '报考车型', '教练', 
#                   '总费用', '状态', '报名费金额', '报名费收取时间', '学费现金金额', '学费现金时间',
#                   '结算完毕', '备注']
#         ws.append(headers)
        
#         # 数据
#         for student in queryset:
#             ws.append([
#                 student.student_id,
#                 student.name,
#                 student.phone,
#                 student.id_card,
#                 student.register_date,
#                 student.get_car_type_display(),
#                 str(student.coach) if student.coach else '',
#                 student.total_fee,
#                 student.get_status_display(),
#                 student.registration_fee,
#                 student.registration_fee_date,
#                 student.tuition_cash,
#                 student.tuition_cash_date,
#                 student.get_settlement_completed_display(),
#                 student.notes
#             ])
        
#         # 响应
#         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = 'attachment; filename=students.xlsx'
#         wb.save(response)
#         return response
    
#     export_to_excel.short_description = "导出选中学员到Excel"

# @admin.register(CoachSettlement)
# class CoachSettlementAdmin(admin.ModelAdmin):
#     list_display = ['coach', 'settlement_month', 'total_amount', 'is_paid', 'created_at']
#     list_filter = ['settlement_month', 'is_paid']
#     search_fields = ['coach__name']
#     filter_horizontal = ['graduated_students']



from django.contrib import admin
from django.utils.html import format_html
from .models import Coach, Student, CoachSettlement
import openpyxl
from django.http import HttpResponse

@admin.register(Coach)
class CoachAdmin(admin.ModelAdmin):
    list_display = ['name', 'phone', 'car_plate', 'created_at']
    search_fields = ['name', 'phone', 'car_plate']
    list_filter = ['created_at']
    actions = ['export_to_excel']
    
    def export_to_excel(self, request, queryset):
        """导出教练数据到Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "教练数据"
        
        # 表头
        headers = ['教练姓名', '联系电话', '车牌号', '创建时间', '更新时间']
        ws.append(headers)
        
        # 数据
        for coach in queryset:
            ws.append([
                coach.name,
                coach.phone,
                coach.car_plate or '',
                coach.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                coach.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        # 响应
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=coaches.xlsx'
        wb.save(response)
        return response
    
    export_to_excel.short_description = "导出选中教练到Excel"

@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    list_display = ['student_id', 'name', 'phone', 'car_type', 'coach', 'status', 'total_fee', 'settlement_completed', 'register_date']
    list_filter = ['status', 'car_type', 'coach', 'register_date', 'settlement_completed']
    search_fields = ['student_id', 'name', 'phone', 'id_card']
    readonly_fields = ['student_id', 'created_at', 'updated_at']
    list_per_page = 20
    actions = ['export_to_excel']
    
    fieldsets = (
        ('基本信息', {
            'fields': ('student_id', 'name', 'phone', 'id_card', 'register_date', 'car_type', 'coach', 'status')
        }),
        ('费用信息', {
            'fields': ('total_fee',)
        }),
        ('资金记录', {
            'fields': (
                ('registration_fee', 'registration_fee_date'),
                ('tuition_cash', 'tuition_cash_date'),
            ),
            'description': '填写报名费和学费收取情况'
        }),
        ('时间信息', {
            'fields': ('transfer_date', 'start_training_date', 'graduation_date')
        }),
        ('银行转账记录', {
            'fields': (
                ('bank_transfer1_amount', 'bank_transfer1_date'),
                ('bank_transfer2_amount', 'bank_transfer2_date'),
                ('bank_transfer3_amount', 'bank_transfer3_date'),
            )
        }),
        ('教练支付与结算', {
            'fields': (
                ('pay_coach_amount', 'pay_coach_date'),
                'settlement_completed',
            )
        }),
        ('其他信息', {
            'fields': ('car_number', 'notes')
        }),
    )
    
    def export_to_excel(self, request, queryset):
        """导出Excel功能"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "学员数据"
        
        # 表头
        headers = ['学员编号', '姓名', '手机号', '身份证号', '报名日期', '报考车型', '教练', 
                  '总费用', '状态', '报名费金额', '报名费收取时间', '学费现金金额', '学费现金时间',
                  '结算完毕', '备注']
        ws.append(headers)
        
        # 数据
        for student in queryset:
            ws.append([
                student.student_id,
                student.name,
                student.phone,
                student.id_card,
                student.register_date,
                student.get_car_type_display(),
                str(student.coach) if student.coach else '',
                student.total_fee,
                student.get_status_display(),
                student.registration_fee,
                student.registration_fee_date,
                student.tuition_cash,
                student.tuition_cash_date,
                student.get_settlement_completed_display(),
                student.notes
            ])
        
        # 响应
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students.xlsx'
        wb.save(response)
        return response
    
    export_to_excel.short_description = "导出选中学员到Excel"

@admin.register(CoachSettlement)
class CoachSettlementAdmin(admin.ModelAdmin):
    list_display = ['coach', 'settlement_month', 'total_amount', 'is_paid', 'created_at']
    list_filter = ['settlement_month', 'is_paid']
    search_fields = ['coach__name']
    filter_horizontal = ['graduated_students']
    actions = ['export_to_excel']
    
    def export_to_excel(self, request, queryset):
        """导出教练结算数据到Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "教练结算数据"
        
        # 表头
        headers = ['教练', '结算月份', '单价', '应结算金额', '是否已支付', '创建时间', '毕业学员数量', '毕业学员名单']
        ws.append(headers)
        
        # 数据
        for settlement in queryset:
            graduated_students = list(settlement.graduated_students.all())
            student_names = '、'.join([student.name for student in graduated_students])
            
            ws.append([
                settlement.coach.name,
                settlement.settlement_month,
                settlement.unit_price,
                settlement.total_amount,
                '是' if settlement.is_paid else '否',
                settlement.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                len(graduated_students),
                student_names
            ])
        
        # 响应
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=coach_settlements.xlsx'
        wb.save(response)
        return response
    
    export_to_excel.short_description = "导出选中教练结算到Excel"
